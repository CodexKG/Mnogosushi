from django.contrib import admin
from rangefilter.filter import DateRangeFilter
from datetime import date, timedelta
from django.utils.translation import gettext as _
from django.db.models import Sum, Count, F
from mptt.admin import MPTTModelAdmin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.billing.models import Billing, BillingProduct, SaleSummary

# Register your models here.
class CustomDateFieldListFilter(admin.DateFieldListFilter):
    def __init__(self, field, request, params, model, model_admin, field_path):
        super().__init__(field, request, params, model, model_admin, field_path)
        self.title = _('Дата')
        self.links = (
            (_('Сегодня'), {
                self.lookup_kwarg_since: str(date.today()),
            }),
            (_('Вчера'), {
                self.lookup_kwarg_since: str(date.today() - timedelta(days=1)),
                self.lookup_kwarg_until: str(date.today() - timedelta(days=1)),
            }),
            (_('Последние 7 дней'), {
                self.lookup_kwarg_since: str(date.today() - timedelta(days=7)),
            }),
            (_('Последние 30 дней'), {
                self.lookup_kwarg_since: str(date.today() - timedelta(days=30)),
            }),
        )


class ProductTabularInline(admin.TabularInline):
    model = BillingProduct
    extra = 0


@admin.register(SaleSummary)
class SaleSummaryAdmin(admin.ModelAdmin):
    change_list_template = 'admin/sale_summary_change_list.html'
    date_hierarchy = 'created'

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(
            request,
            extra_context=extra_context,
        )

        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        metrics = {
            'title': F('billing_products__product__title'),  # Замените 'billing__title' на фактический путь к полю 'title' в модели BillingProduct
            'total': F('billing_products__quantity'),
            'total_sales': Sum('billing_products__price'),
            'payment_method': F('payment_method'),
        }

        response.context_data['summary'] = list(
            qs.values('billing_products').annotate(**metrics).order_by('-created')
        )
        print(response)
        print(metrics)

        ################################
        total_metrics = {
            'title': Sum('billing_products__product__title'),  # Замените 'billing__title' на фактический путь к полю 'title' в модели BillingProduct
            'total': Sum('billing_products__quantity'),
            'total_sales': Sum('billing_products__price'),
            'payment_method': Count('payment_method'),
        }

        response.context_data['summary_total'] = dict(
            qs.aggregate(**total_metrics)
        )

        return response
    
    def get_next_in_date_hierarchy(request, date_hierarchy):
        if date_hierarchy + '__day' in request.GET:
            return 'hour'

        if date_hierarchy + '__month' in request.GET:
            return 'day'

        if date_hierarchy + '__year' in request.GET:
            return 'week'

        return 'month'

def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="billings.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Billings"

    columns = ['ID', 'Итоговая цена', 'Цена доставки', 'Адрес', 'Номер телефона', 'Оплата', 'Код оплаты', 'Статус', 'Создан']
    ws.append(columns)

    # Track the maximum length of each column to set the column width later
    column_widths = [len(column) for column in columns]

    for obj in queryset:
        # Convert Boolean status to a more user-friendly format
        status_display = "Оплачено" if obj.status else "Не оплачено"
        
        row = [
            obj.id, obj.total_price, obj.delivery_price, obj.address, obj.phone,
            obj.payment_method, obj.payment_code, status_display, obj.created.strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row)
        for i, cell in enumerate(row):
            # Update the maximum length found in each column
            column_widths[i] = max(column_widths[i], len(str(cell)))

    # Set the column widths based on the maximum length found, adding a small buffer
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width + 2  # Adding a buffer of 2

    wb.save(response)
    return response

export_to_excel.short_description = "Export to Excel"

@admin.register(Billing)
class BillingAdmin(MPTTModelAdmin):
    list_display = ('id', 'total_price', 'address', 'payment_method', 'phone', 'delivery_price', 'billing_receipt_type', 'payment_code', 'created', 'status')
    search_fields = ('id', 'total_price', 'address', 'payment_method', 'phone', 'delivery_price', 'billing_receipt_type', 'payment_code', 'created', 'status')
    ordering = ('-created', )
    inlines = [ProductTabularInline]
    list_filter = (('created', DateRangeFilter), ('created', CustomDateFieldListFilter),)  # Добавляем фильтр по дате
    actions = [export_to_excel]


@admin.register(BillingProduct)
class BillingProductAdmin(admin.ModelAdmin):
    list_display = ('billing', 'product', 'quantity', 'price', 'status')
    list_filter = (('created', DateRangeFilter), ('created', CustomDateFieldListFilter),)  # Добавляем фильтр по датее